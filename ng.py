import time
import ray
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook



def stress_function(num):
    return sum([i * j * k for i in range(num) for j in range(i) for k in range(j)])

@ray.remote
def mega_work(start, end):
    return [stress_function(x) for x in range(start, end)]

if __name__ == "__main__":
    ray.init(address='auto')

    inp = int(input("Enter a number divided by 25: "))
    assert inp%25 == 0, "Number should be divided by 50"
    czas = []

    try:
        wb = Workbook()
    except:
        wb = load_workbook("excel_data.xlsx")
    sheet = wb.active

    time_start = time.time()
    for i in range(0, inp, 25):
        start = time.time()
        futures = [(mega_work.remote(x*10, (x+1)*10)) for x in range(i)]
        ray.get(futures)
        czas.append(time.time() - start)
    print(time.time() - time_start)

    x = int(inp / 25)
    for i in range(1, x):
        sheet.cell(row=i, column=1).value = i*25
        sheet.cell(row=i, column=2).value = czas[i]

    wb.save("excel_data.xlsx")

    dev_x = [i for i in range(1, inp, 25)]
    dev_y = [czas[i] for i in range(x)]

    plt.xlabel("Number of iterations")
    plt.ylabel("Time in seconds")
    plt.title("Time of compilation by number of iterations")

    plt.plot(dev_x, dev_y)

    plt.legend(["Stress Function"])

    plt.show(block=True)
